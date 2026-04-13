"""
Task 1.2: Impute missing river water levels using each river's own past values.

Strategy:
  1. Load the river data, auto-detect the header row, and normalize columns.
  2. Drop entirely-empty rows (placeholders for large gaps).
  3. Reindex to a continuous monthly grid so all gaps become explicit NaNs.
  4. Impute using a two-step approach:
     a) Same-month climatological mean  – fills NaN with the average of the
        same calendar month across all available years (captures seasonality).
     b) Linear interpolation – for any remaining isolated gaps.
  5. Save the imputed data back to new Excel files.
"""

import pandas as pd
import numpy as np
import os
import warnings
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

# ============================================================================
# SETUP: Define input/output directories and find all river Excel files
# ============================================================================

# Path to the folder containing the original river Excel files
DATA_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "Donnees Hydro et Meteo Mensuelles_IGEBU",
)

# Path to the folder where imputed results will be saved
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "imputed_output")
os.makedirs(OUTPUT_DIR, exist_ok=True)  # Create it if it doesn't exist

# Get a sorted list of all river files (files with "Riviere" in the name)
river_files = sorted(
    f for f in os.listdir(DATA_DIR) if "Riviere" in f and f.endswith(".xlsx")
)


# ============================================================================
# HELPER FUNCTION 1: Auto-detect which row in the Excel file is the header
# ============================================================================
# Each Excel file has a different structure — some have the header on row 1,
# some on row 3, some on row 4. This function scans the first 10 rows and
# finds the one that contains "Year" or "YY" (indicating it's the header).
def find_header_row(filepath):
    """Find the row index (0-based) that contains the header."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        vals = [str(v).strip().lower() if v else "" for v in row]
        if "year" in vals or "yy" in vals:
            wb.close()
            return i
    wb.close()
    return 0


# ============================================================================
# HELPER FUNCTION 2: Normalize column names to a standard set
# ============================================================================
# Different files use different column names (e.g., "Alt" vs "Elev",
# "YY" vs "Year", "haut moy" vs "haut.moy" vs "Haut.moy").
# This function renames them all to one consistent set.
def normalize_columns(df):
    """Normalize column names to a standard set."""
    col_map = {}
    for c in df.columns:
        cl = str(c).strip().lower()
        if cl in ("station_name", "name"):
            col_map[c] = "Station_Name"
        elif cl == "stationid":
            col_map[c] = "StationID"
        elif cl == "lat":
            col_map[c] = "Lat"
        elif cl == "lon":
            col_map[c] = "Lon"
        elif cl in ("elev", "alt"):
            col_map[c] = "Elev"
        elif cl in ("year", "yy"):
            col_map[c] = "Year"
        elif cl in ("month", "mm"):
            col_map[c] = "Month"
        elif "haut" in cl and "moy" in cl:
            col_map[c] = "haut_moy"  # This is our target variable (water level)
    df = df.rename(columns=col_map)
    return df


# ============================================================================
# MAIN LOOP: Process each river file one by one
# ============================================================================

print(f"Found {len(river_files)} river files.\n")
print("=" * 80)

for fname in river_files:
    path = os.path.join(DATA_DIR, fname)

    # Extract the river name from the filename (e.g., "Buzimba" from the full name)
    river_name = (
        fname.replace("Niveaux d'eau de la Riviere ", "").replace(".xlsx", "")
    )

    # -----------------------------------------------------------------------
    # STEP 1: Load the Excel file
    # -----------------------------------------------------------------------
    # Auto-detect which row is the header, then read the file with pandas
    header_row = find_header_row(path)
    df = pd.read_excel(path, header=header_row)

    # Rename all columns to our standard names
    df = normalize_columns(df)

    # The Mpanda file has an extra "StationID" column — drop it
    if "StationID" in df.columns:
        df = df.drop(columns=["StationID"])

    # Check that all required columns exist
    required = ["Station_Name", "Lat", "Lon", "Elev", "Year", "Month", "haut_moy"]
    for col in required:
        if col not in df.columns:
            print(f"\n[{river_name}]  WARNING: missing column '{col}'. Columns: {list(df.columns)}")
            break
    else:
        pass  # All columns present, continue

    # Keep only the columns we need
    df = df[required]

    # -----------------------------------------------------------------------
    # STEP 2: Clean the data — remove empty rows and invalid entries
    # -----------------------------------------------------------------------

    # Some rows in the Excel are completely empty (all None) — remove them
    df = df.dropna(how="all").reset_index(drop=True)

    # Some rows have no Year/Month — these are placeholders for gaps, remove them
    df = df.dropna(subset=["Year", "Month"]).reset_index(drop=True)

    # Make sure Year and Month are integers (some files have them as strings)
    df["Year"] = pd.to_numeric(df["Year"], errors="coerce")
    df["Month"] = pd.to_numeric(df["Month"], errors="coerce")
    df = df.dropna(subset=["Year", "Month"]).reset_index(drop=True)
    df["Year"] = df["Year"].astype(int)
    df["Month"] = df["Month"].astype(int)

    # -----------------------------------------------------------------------
    # STEP 3: Create a proper date column and sort chronologically
    # -----------------------------------------------------------------------

    # Combine Year + Month into a proper datetime (day is always 1st of month)
    df["Date"] = pd.to_datetime(df[["Year", "Month"]].assign(Day=1))

    # Sort by date and remove any duplicate months
    df = df.sort_values("Date").drop_duplicates(subset=["Date"]).reset_index(drop=True)

    # Fill in station metadata (name, lat, lon, elevation) for any NaN rows
    meta_cols = ["Station_Name", "Lat", "Lon", "Elev"]
    for col in meta_cols:
        df[col] = df[col].ffill().bfill()

    # Save metadata values for later (to fill in newly created rows)
    station_name = df["Station_Name"].iloc[0]
    lat = df["Lat"].iloc[0]
    lon = df["Lon"].iloc[0]
    elev = df["Elev"].iloc[0]

    # -----------------------------------------------------------------------
    # STEP 4: Reindex to a COMPLETE monthly grid
    # -----------------------------------------------------------------------
    # The original data might jump from e.g. 1982-08 to 1984-02, skipping
    # 17 months. Here we create a row for EVERY month in the range, so those
    # gaps become explicit NaN values that we can then fill.

    full_range = pd.date_range(start=df["Date"].min(), end=df["Date"].max(), freq="MS")
    df = df.set_index("Date").reindex(full_range)  # NaN rows appear for missing months
    df.index.name = "Date"

    # Fill metadata for the newly created rows
    df["Station_Name"] = station_name
    df["Lat"] = lat
    df["Lon"] = lon
    df["Elev"] = elev
    df["Year"] = df.index.year
    df["Month"] = df.index.month

    # -----------------------------------------------------------------------
    # STEP 5: Check how many values are missing
    # -----------------------------------------------------------------------

    n_missing_before = df["haut_moy"].isna().sum()

    # If no missing values, skip this river
    if n_missing_before == 0:
        print(f"\n[{river_name}]  No missing values. Skipping.")
        continue

    # Print summary info
    print(f"\n[{river_name}]")
    print(
        f"  Period        : {df.index.min().strftime('%Y-%m')} to {df.index.max().strftime('%Y-%m')}"
    )
    print(f"  Total months  : {len(df)}")
    print(f"  Missing before: {n_missing_before}")

    # Show which specific months are missing
    missing_mask = df["haut_moy"].isna()
    missing_dates = df.index[missing_mask]
    print(f"  Missing dates : {[d.strftime('%Y-%m') for d in missing_dates]}")

    # -----------------------------------------------------------------------
    # STEP 6 (IMPUTATION): Same-month climatological mean
    # -----------------------------------------------------------------------
    # This is the CORE of the imputation.
    #
    # For each calendar month (Jan=1, Feb=2, ..., Dec=12), we compute the
    # AVERAGE water level across ALL years where that month has data.
    #
    # Example for Buzimba, month=4 (April):
    #   April 1981: 0.648
    #   April 1982: 0.672
    #   April 1984: 0.619
    #   ...
    #   April 2013: 1.250
    #   → Mean of all Aprils = 1.1435
    #
    # Then if April 2011 is missing, we fill it with 1.1435.
    # This captures the SEASONAL PATTERN of the river (high in rainy months,
    # low in dry months).

    monthly_means = df.groupby("Month")["haut_moy"].mean()

    # Print the climatology table for this river
    print(f"\n  Monthly climatology (mean water level by calendar month):")
    for m in range(1, 13):
        val = monthly_means.get(m, np.nan)
        bar = "█" * int(val * 20) if not np.isnan(val) else ""
        print(f"    Month {m:2d}: {val:.4f}  {bar}")

    # Create a copy of the original column to preserve it
    df["haut_moy_imputed"] = df["haut_moy"].copy()

    # Loop through every row that has a missing value and fill it
    for idx in df.index[df["haut_moy_imputed"].isna()]:
        month = idx.month  # Which calendar month is this? (1-12)
        climatological_mean = monthly_means.get(month, np.nan)  # Get the average for that month
        if not np.isnan(climatological_mean):
            df.loc[idx, "haut_moy_imputed"] = climatological_mean  # Fill it in!
            print(
                f"  -> Imputed {idx.strftime('%Y-%m')} with month-{month} mean = {climatological_mean:.4f}"
            )

    # -----------------------------------------------------------------------
    # STEP 7 (FALLBACK): Linear interpolation for any remaining gaps
    # -----------------------------------------------------------------------
    # If there are still NaN values (e.g., a month where NO year had data),
    # we use linear interpolation: draw a straight line between the nearest
    # known values on each side and pick the point in between.

    remaining_missing = df["haut_moy_imputed"].isna().sum()
    if remaining_missing > 0:
        df["haut_moy_imputed"] = df["haut_moy_imputed"].interpolate(method="linear")
        print(
            f"  -> Used linear interpolation for {remaining_missing} remaining gap(s)"
        )

    # Verify: how many are still missing? (should be 0)
    n_missing_after = df["haut_moy_imputed"].isna().sum()
    print(f"  Missing after : {n_missing_after}")

    # -----------------------------------------------------------------------
    # STEP 8: Save the result to a new Excel file
    # -----------------------------------------------------------------------
    # The output has BOTH the original column (with NaNs) and the imputed
    # column (all filled), so you can always compare them.

    out_df = df[
        [
            "Station_Name",
            "Lat",
            "Lon",
            "Elev",
            "Year",
            "Month",
            "haut_moy",          # Original values (NaN where missing)
            "haut_moy_imputed",  # Imputed values (all filled)
        ]
    ].copy()
    out_df = out_df.rename(columns={"haut_moy": "haut_moy_original"})
    out_df = out_df.reset_index().rename(columns={"index": "Date"})

    out_path = os.path.join(OUTPUT_DIR, f"{river_name}_imputed.xlsx")
    out_df.to_excel(out_path, index=False)
    print(f"  Saved to: {out_path}")

print("\n" + "=" * 80)
print("Done! All imputed files are in:", OUTPUT_DIR)
